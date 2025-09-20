import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

from app.config import settings
from app.utils.logger import get_logger
from app.models.requirement import Requirement, RequirementType, Priority, Status
from app.models.rtm import RTMOutput

logger = get_logger(__name__)

class RTMOutputGenerator:
    """
    Generates comprehensive 3-sheet RTM output:
    1. Detailed Focus Sheet Analysis
    2. Complete RTM (All Sheets)  
    3. Summary Statistics
    """
    
    def __init__(self):
        self.logger = logger
        
        # Excel styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_complete_rtm(self, focus_sheet_analysis: List[Dict], 
                            all_sheets_analysis: Dict[str, List[Dict]], 
                            source_file_info: Dict,
                            focus_sheet_name: str) -> RTMOutput:
        """
        Generate complete 3-sheet RTM output
        """
        try:
            start_time = datetime.now()
            self.logger.info("🚀 Starting RTM generation with 3-sheet structure")
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            source_name = Path(source_file_info.get('file_name', 'requirements')).stem
            output_filename = f"RTM_{source_name}_{timestamp}.xlsx"
            output_path = Path(settings.OUTPUT_DIR) / output_filename
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Sheet 1: Detailed Focus Sheet Analysis
            self.logger.info(f"📋 Creating Sheet 1: Detailed Analysis - {focus_sheet_name}")
            focus_sheet = workbook.create_sheet(f"Detailed Analysis - {focus_sheet_name}")
            self._create_focus_sheet_rtm(focus_sheet, focus_sheet_analysis, focus_sheet_name)
            
            # Sheet 2: Complete RTM (All Sheets in Order)
            self.logger.info("📋 Creating Sheet 2: Complete RTM - All Sheets")
            complete_sheet = workbook.create_sheet("Complete RTM - All Sheets")
            self._create_complete_rtm_sheet(complete_sheet, all_sheets_analysis, source_file_info)
            
            # Sheet 3: Summary Statistics
            self.logger.info("📋 Creating Sheet 3: Summary Statistics")
            summary_sheet = workbook.create_sheet("Summary Statistics")
            summary_stats = self._create_summary_sheet(summary_sheet, all_sheets_analysis, 
                                                     focus_sheet_name, focus_sheet_analysis)
            
            # Save workbook
            workbook.save(output_path)
            
            # Calculate processing time
            processing_time = (datetime.now() - start_time).total_seconds()
            
            # Count total requirements with detailed logging
            total_requirements = 0
            detailed_count = {}
            
            for sheet_name, requirements in all_sheets_analysis.items():
                sheet_count = len(requirements)
                total_requirements += sheet_count
                detailed_count[sheet_name] = sheet_count
                
                # Additional analysis for enhanced validation
                validated_reqs = len([req for req in requirements if not req.get('is_edge_case', False)])
                edge_cases = sheet_count - validated_reqs
                
                self.logger.info(f"📊 {sheet_name}: {sheet_count} total ({validated_reqs} validated, {edge_cases} edge cases)")
            
            self.logger.info(f"📈 Total requirements across all sheets: {total_requirements}")
            self.logger.info(f"📋 Sheet breakdown: {detailed_count}")
            
            # Create RTM output object
            rtm_output = RTMOutput(
                file_path=str(output_path),
                requirements_count=total_requirements,
                summary_statistics=summary_stats,
                processing_time=processing_time,
                source_file_name=source_file_info.get('file_name', 'Unknown'),
                generated_at=start_time
            )
            
            self.logger.info(f"✅ RTM generation completed in {processing_time:.2f} seconds")
            self.logger.info(f"📁 Output file: {output_filename}")
            
            return rtm_output
        
        except Exception as e:
            self.logger.error(f"Error creating RTM: {str(e)}")
            raise

    def _safe_excel_value(self, value):
        """
        Safely convert any value to Excel-compatible format
        """
        if value is None:
            return ""
        elif isinstance(value, list):
            # Convert list to comma-separated string
            return ", ".join(str(item) for item in value if item is not None)
        elif isinstance(value, dict):
            # Convert dict to readable string
            return str(value)
        elif isinstance(value, (int, float, str, bool)):
            return value
        else:
            # Convert other types to string
            return str(value)
    
    def _create_focus_sheet_rtm(self, worksheet, focus_analysis: List[Dict], sheet_name: str):
        """
        Create detailed RTM for focus sheet with comprehensive analysis
        """
        # Define columns for detailed analysis
        columns = [
            'Requirement ID', 'Requirement Description', 'Source', 
            'Requirement Type', 'Priority', 'Priority Reasoning',
            'Status', 'Related Deliverables', 'Test Case ID', 
            'Test Case Suggestions', 'Comments', 'Analysis Confidence',
            'Original ID', 'AI Analysis Used'
        ]
        
        # Add headers
        for col_idx, column in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = self.border
        
        # Add data rows
        req_counter = 1
        for row_idx, req in enumerate(focus_analysis, 2):
            # Generate sequential IDs
            req_id = f"{settings.REQUIREMENT_ID_PREFIX}-{req_counter:03d}"
            test_id = f"{settings.TEST_CASE_ID_PREFIX}-{req_counter:03d}"
            
            row_data = [
                req_id,
                req.get('original_requirement', ''),
                req.get('source', ''),
                req.get('requirement_type', 'Functional'),
                req.get('priority', 'Medium'),
                req.get('priority_reasoning', ''),
                'Not Tested',  # Default status
                req.get('related_deliverables', ''),
                test_id,
                '\n'.join(req.get('test_case_suggestions', [])),
                req.get('comments', ''),
                req.get('analysis_confidence', 1.0),
                req.get('original_id', ''),
                'AI' if not req.get('fallback_analysis', False) else 'Rule-based'
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = self._safe_excel_value(value)
                cell.border = self.border
                
                # Wrap text for long content
                if col_idx in [2, 6, 10, 11]:  # Description, reasoning, test cases, comments
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                
            req_counter += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet, columns)
        
        # Add sheet information
        info_row = len(focus_analysis) + 3
        worksheet.cell(row=info_row, column=1).value = "Sheet Information:"
        worksheet.cell(row=info_row, column=1).font = Font(bold=True)
        
        info_data = [
            f"Focus Sheet: {sheet_name}",
            f"Total Requirements: {len(focus_analysis)}",
            f"Analysis Type: Detailed (Priority Sheet)",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        
        for i, info in enumerate(info_data):
            worksheet.cell(row=info_row + i + 1, column=1).value = info
    
    def _create_complete_rtm_sheet(self, worksheet, all_sheets_analysis: Dict[str, List[Dict]], 
                                 source_file_info: Dict):
        """
        Create complete RTM with all sheets in original file order
        """
        # Define columns for complete RTM
        columns = [
            'Requirement ID', 'Requirement Description', 'Source Sheet', 
            'Source Reference', 'Requirement Type', 'Priority', 
            'Status', 'Related Deliverables', 'Test Case ID', 
            'Comments', 'Original ID'
        ]
        
        # Add headers
        for col_idx, column in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = self.border
        
        # Add data from all sheets in original order
        req_counter = 1
        current_row = 2
        
        # Process sheets in original file order
        original_sheet_order = source_file_info.get('sheet_names', [])
        
        for sheet_name in original_sheet_order:
            if sheet_name not in all_sheets_analysis:
                continue
                
            sheet_requirements = all_sheets_analysis[sheet_name]
            
            if not sheet_requirements:
                continue
            
            # Add sheet separator
            if current_row > 2:  # Not the first sheet
                separator_row = current_row
                worksheet.cell(row=separator_row, column=1).value = f"--- {sheet_name} ---"
                worksheet.cell(row=separator_row, column=1).font = Font(bold=True, italic=True)
                
                # Merge cells for separator
                worksheet.merge_cells(f"A{separator_row}:K{separator_row}")
                current_row += 1
            
            # Add requirements from this sheet
            for req in sheet_requirements:
                req_id = f"{settings.REQUIREMENT_ID_PREFIX}-{req_counter:03d}"
                test_id = f"{settings.TEST_CASE_ID_PREFIX}-{req_counter:03d}"
                
                row_data = [
                    req_id,
                    req.get('original_requirement', ''),
                    sheet_name,
                    req.get('source', ''),
                    req.get('requirement_type', 'Functional'),
                    req.get('priority', 'Medium'),
                    'Not Tested',
                    req.get('related_deliverables', ''),
                    test_id,
                    req.get('comments', ''),
                    req.get('original_id', '')
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=current_row, column=col_idx)
                    cell.value = self._safe_excel_value(value)
                    cell.border = self.border
                    
                    # Wrap text for long content
                    if col_idx in [2, 8, 10]:  # Description, deliverables, comments
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                current_row += 1
                req_counter += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet, columns)
    
    def _create_summary_sheet(self, worksheet, all_sheets_analysis: Dict[str, List[Dict]], 
                            focus_sheet_name: str, focus_analysis: List[Dict]) -> Dict:
        """
        Create summary statistics sheet
        """
        # Calculate statistics
        total_requirements = sum(len(reqs) for reqs in all_sheets_analysis.values())
        total_sheets = len(all_sheets_analysis)
        focus_requirements = len(focus_analysis)
        
        # Count by type and priority
        type_counts = {}
        priority_counts = {}
        sheet_counts = {}
        ai_analysis_count = 0
        fallback_count = 0
        
        for sheet_name, requirements in all_sheets_analysis.items():
            sheet_counts[sheet_name] = len(requirements)
            
            for req in requirements:
                # Count by type
                req_type = req.get('requirement_type', 'Unknown')
                type_counts[req_type] = type_counts.get(req_type, 0) + 1
                
                # Count by priority
                priority = req.get('priority', 'Unknown')
                priority_counts[priority] = priority_counts.get(priority, 0) + 1
                
                # Count AI vs fallback
                if req.get('fallback_analysis', False):
                    fallback_count += 1
                else:
                    ai_analysis_count += 1
        
        # Create summary data structure
        summary_stats = {
            'total_requirements': total_requirements,
            'total_sheets': total_sheets,
            'focus_sheet': focus_sheet_name,
            'focus_requirements': focus_requirements,
            'by_type': type_counts,
            'by_priority': priority_counts,
            'by_sheet': sheet_counts,
            'ai_analysis_count': ai_analysis_count,
            'fallback_count': fallback_count
        }
        
        # Write to worksheet
        row = 1
        
        # Title
        worksheet.cell(row=row, column=1).value = "Requirements Traceability Matrix - Summary"
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=16)
        row += 3
        
        # General statistics
        worksheet.cell(row=row, column=1).value = "General Statistics"
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        
        general_stats = [
            ("Total Requirements", total_requirements),
            ("Total Sheets Processed", total_sheets),
            ("Focus Sheet", focus_sheet_name),
            ("Focus Sheet Requirements", focus_requirements),
            ("AI Analysis Used", ai_analysis_count),
            ("Rule-based Fallback Used", fallback_count),
            ("Processing Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for stat_name, stat_value in general_stats:
            worksheet.cell(row=row, column=1).value = stat_name
            worksheet.cell(row=row, column=2).value = self._safe_excel_value(stat_value)
            worksheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Requirements by Type
        worksheet.cell(row=row, column=1).value = "Requirements by Type"
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        
        for req_type, count in sorted(type_counts.items()):
            worksheet.cell(row=row, column=1).value = req_type
            worksheet.cell(row=row, column=2).value = count
            worksheet.cell(row=row, column=3).value = f"{(count/total_requirements)*100:.1f}%"
            row += 1
        
        row += 2
        
        # Requirements by Priority
        worksheet.cell(row=row, column=1).value = "Requirements by Priority"
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        
        priority_order = ['High', 'Medium', 'Low']
        for priority in priority_order:
            if priority in priority_counts:
                count = priority_counts[priority]
                worksheet.cell(row=row, column=1).value = priority
                worksheet.cell(row=row, column=2).value = count
                worksheet.cell(row=row, column=3).value = f"{(count/total_requirements)*100:.1f}%"
                row += 1
        
        row += 2
        
        # Requirements by Sheet
        worksheet.cell(row=row, column=1).value = "Requirements by Sheet"
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        
        for sheet_name, count in sheet_counts.items():
            worksheet.cell(row=row, column=1).value = sheet_name
            worksheet.cell(row=row, column=2).value = count
            worksheet.cell(row=row, column=3).value = f"{(count/total_requirements)*100:.1f}%"
            # Mark focus sheet
            if sheet_name == focus_sheet_name:
                worksheet.cell(row=row, column=4).value = "FOCUS SHEET"
                worksheet.cell(row=row, column=4).font = Font(bold=True, color="FF0000")
            row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(worksheet, ["Metric", "Value", "Percentage", "Notes"])
        
        return summary_stats
    
    def _auto_adjust_columns(self, worksheet, columns: List[str]):
        """
        Auto-adjust column widths based on content
        """
        for col_idx, column in enumerate(columns, 1):
            max_length = len(column)  # Start with header length
            
            # Check data length
            for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            
            # Set column width (with some padding)
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width