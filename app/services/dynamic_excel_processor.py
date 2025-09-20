import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import List, Dict, Any, Optional, Tuple
import re
from pathlib import Path

from app.config import settings
from app.utils.logger import get_logger
from app.utils.exceptions import ExcelProcessingError, FileHandlingError
from app.services.universal_requirement_validator import UniversalRequirementValidator

logger = get_logger(__name__)

class DynamicExcelProcessor:
    """
    Enhanced Excel processor that can dynamically handle various Excel file structures
    and identify requirements across different formats and layouts.
    Integrates with UniversalRequirementValidator for comprehensive requirement detection.
    """
    
    def __init__(self):
        self.logger = logger
        
        # Initialize the universal validator
        self.validator = UniversalRequirementValidator()
        
        # Patterns for identifying requirement columns (legacy - validator handles this better)
        self.requirement_patterns = [
            r'req', r'requirement', r'spec', r'specification', r'need', r'shall', 
            r'must', r'should', r'will', r'description', r'function', r'feature',
            r'capability', r'objective', r'goal', r'criteria'
        ]
        
        # Patterns for identifying ID columns
        self.id_patterns = [
            r'id', r'number', r'no', r'ref', r'reference', r'code', r'identifier'
        ]
        
        # Patterns for priority/status columns
        self.priority_patterns = [
            r'priority', r'importance', r'criticality', r'level', r'status'
        ]
        
    def load_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Load Excel file and analyze its structure dynamically using enhanced validator
        """
        try:
            self.logger.info(f"📂 Loading Excel file: {Path(file_path).name}")
            
            # Load with openpyxl to handle merged cells
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Also load with pandas for easier data manipulation
            excel_sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            
            file_info = {
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'sheet_names': list(excel_sheets.keys()),
                'total_sheets': len(excel_sheets),
                'workbook': workbook,
                'pandas_sheets': excel_sheets,
                'sheets_analysis': {}
            }
            
            self.logger.info(f"✅ Loaded Excel file with {len(excel_sheets)} sheets: {', '.join(excel_sheets.keys())}")
            
            # Analyze each sheet
            for sheet_name in excel_sheets.keys():
                sheet_analysis = self._analyze_sheet_structure(
                    sheet_name, excel_sheets[sheet_name], workbook[sheet_name]
                )
                file_info['sheets_analysis'][sheet_name] = sheet_analysis
            
            return file_info
            
        except Exception as e:
            self.logger.error(f"Error loading Excel file {file_path}: {str(e)}")
            raise ExcelProcessingError(f"Failed to load Excel file: {str(e)}")
    
    def _analyze_sheet_structure(self, sheet_name: str, df: pd.DataFrame, ws) -> Dict[str, Any]:
        """
        Analyze individual sheet structure and detect columns dynamically
        """
        try:
            self.logger.info(f"🔍 Analyzing sheet: '{sheet_name}'")
            
            # Basic sheet info
            analysis = {
                'sheet_name': sheet_name,
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'column_names': list(df.columns),
                'detected_columns': {},
                'merged_cells': [],
                'potential_requirement_columns': [],
                'has_requirements': False,
                'confidence_score': 0.0
            }
            
            # Detect merged cells
            analysis['merged_cells'] = self._detect_merged_cells(ws)
            
            # Analyze columns to detect their purposes
            for col_idx, column_name in enumerate(df.columns):
                column_analysis = self._analyze_column(df, column_name, col_idx)
                analysis['detected_columns'][column_name] = column_analysis
                
                # Check if this looks like a requirements column
                if column_analysis['is_requirement_column']:
                    analysis['potential_requirement_columns'].append(column_name)
            
            # Determine if sheet contains requirements
            analysis['has_requirements'] = len(analysis['potential_requirement_columns']) > 0
            analysis['confidence_score'] = self._calculate_sheet_confidence(analysis)
            
            self.logger.info(f"   📊 Sheet analysis: {analysis['total_rows']} rows, {analysis['total_columns']} columns")
            self.logger.info(f"   🎯 Requirement columns found: {len(analysis['potential_requirement_columns'])}")
            self.logger.info(f"   📈 Confidence score: {analysis['confidence_score']:.2f}")
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"Error analyzing sheet '{sheet_name}': {str(e)}")
            return {
                'sheet_name': sheet_name,
                'error': str(e),
                'has_requirements': False,
                'confidence_score': 0.0
            }
    
    def _detect_merged_cells(self, worksheet) -> List[Dict]:
        """
        Detect and record merged cells for proper content extraction
        """
        merged_cells = []
        
        try:
            for merged_range in worksheet.merged_cells.ranges:
                merged_cells.append({
                    'range': str(merged_range),
                    'start_row': merged_range.min_row,
                    'end_row': merged_range.max_row,
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col,
                    'value': worksheet.cell(merged_range.min_row, merged_range.min_col).value
                })
        except Exception as e:
            self.logger.warning(f"Could not detect merged cells: {str(e)}")
            
        return merged_cells
    
    def _analyze_column(self, df: pd.DataFrame, column_name: str, col_idx: int) -> Dict[str, Any]:
        """
        Analyze individual column to determine its purpose
        """
        column_analysis = {
            'column_name': column_name,
            'column_index': col_idx,
            'data_type': str(df[column_name].dtype),
            'non_null_count': df[column_name].count(),
            'null_count': df[column_name].isnull().sum(),
            'unique_values': df[column_name].nunique(),
            'is_requirement_column': False,
            'is_id_column': False,
            'is_priority_column': False,
            'column_type': 'unknown',
            'confidence': 0.0,
            'sample_values': []
        }
        
        # Get sample non-null values
        non_null_values = df[column_name].dropna()
        if len(non_null_values) > 0:
            sample_size = min(5, len(non_null_values))
            column_analysis['sample_values'] = non_null_values.head(sample_size).tolist()
        
        # Analyze column name patterns
        column_name_lower = column_name.lower()
        
        # Check for requirement column patterns
        requirement_score = 0
        for pattern in self.requirement_patterns:
            if re.search(pattern, column_name_lower):
                requirement_score += 1
        
        # Check content patterns for requirements
        content_requirement_score = self._analyze_content_for_requirements(non_null_values)
        
        total_req_score = requirement_score + content_requirement_score
        if total_req_score >= 1:
            column_analysis['is_requirement_column'] = True
            column_analysis['column_type'] = 'requirement'
            column_analysis['confidence'] = min(1.0, total_req_score / 3)
        
        # Check for ID column patterns
        id_score = 0
        for pattern in self.id_patterns:
            if re.search(pattern, column_name_lower):
                id_score += 1
        
        if id_score > 0 or self._looks_like_id_content(non_null_values):
            column_analysis['is_id_column'] = True
            if column_analysis['column_type'] == 'unknown':
                column_analysis['column_type'] = 'id'
                column_analysis['confidence'] = 0.8
        
        # Check for priority column patterns
        priority_score = 0
        for pattern in self.priority_patterns:
            if re.search(pattern, column_name_lower):
                priority_score += 1
        
        if priority_score > 0:
            column_analysis['is_priority_column'] = True
            if column_analysis['column_type'] == 'unknown':
                column_analysis['column_type'] = 'priority'
                column_analysis['confidence'] = 0.7
        
        return column_analysis
    
    def _analyze_content_for_requirements(self, values: pd.Series) -> int:
        """
        Analyze content to see if it looks like requirements
        """
        if len(values) == 0:
            return 0
        
        score = 0
        sample_values = values.head(10)
        
        for value in sample_values:
            if pd.isna(value):
                continue
                
            value_str = str(value).lower()
            
            # Look for requirement-like language
            if any(word in value_str for word in ['shall', 'must', 'should', 'will', 'can', 'may']):
                score += 0.5
            
            # Look for descriptive content (longer text)
            if len(value_str) > 20:
                score += 0.3
            
            # Look for requirement numbering patterns
            if re.search(r'\d+\.\d+|\d+\)', value_str):
                score += 0.2
        
        return int(score)
    
    def _looks_like_id_content(self, values: pd.Series) -> bool:
        """
        Check if content looks like IDs
        """
        if len(values) == 0:
            return False
        
        sample_values = values.head(10)
        id_like_count = 0
        
        for value in sample_values:
            if pd.isna(value):
                continue
                
            value_str = str(value)
            
            # Check for common ID patterns
            if (re.match(r'^[A-Z]{2,5}-?\d+$', value_str) or  # REQ-001, TC-123
                re.match(r'^\d+(\.\d+)*$', value_str) or      # 1.1, 1.2.3
                re.match(r'^[A-Z]+\d+$', value_str)):         # REQ001, TC123
                id_like_count += 1
        
        return id_like_count >= len(sample_values) * 0.7
    
    def _calculate_sheet_confidence(self, analysis: Dict) -> float:
        """
        Calculate confidence score for sheet containing requirements
        """
        score = 0.0
        
        # Base score for having requirement columns
        if analysis['potential_requirement_columns']:
            score += 0.5
        
        # Bonus for multiple requirement columns
        if len(analysis['potential_requirement_columns']) > 1:
            score += 0.2
        
        # Bonus for having ID columns
        id_columns = [col for col, data in analysis['detected_columns'].items() 
                     if data['is_id_column']]
        if id_columns:
            score += 0.2
        
        # Bonus for reasonable amount of data
        if analysis['total_rows'] > 5:
            score += 0.1
        
        return min(1.0, score)
    
    def extract_requirements_from_sheet(self, sheet_name: str, file_info: Dict, 
                                      preserve_original_ids: bool = True) -> List[Dict]:
        """
        Extract requirements from a specific sheet using comprehensive validation
        Ensures ALL possible requirements are captured regardless of column names or formatting
        """
        try:
            self.logger.info(f"📋 Extracting requirements from sheet: '{sheet_name}' (Enhanced Mode)")
            
            if sheet_name not in file_info['pandas_sheets']:
                self.logger.error(f"Sheet '{sheet_name}' not found in file data")
                return []
            
            # Get the pandas DataFrame for this sheet
            df = file_info['pandas_sheets'][sheet_name]
            
            # Use the universal validator for comprehensive analysis
            validation_results = self.validator.validate_excel_requirements(df, sheet_name)
            
            # Convert validated requirements to our standard format
            requirements = []
            
            for candidate in validation_results['validated_requirements']:
                requirement = {
                    'description': candidate.content,
                    'source': f"{sheet_name}!{candidate.source_column}{candidate.row_index + 1}",
                    'sheet_name': sheet_name,
                    'row_number': candidate.metadata['original_row'] + 1,
                    'column_name': candidate.source_column,
                    'original_id': self._extract_id_from_metadata(candidate, df),
                    'confidence_score': candidate.confidence_score,
                    'category': candidate.category,
                    'additional_info': self._extract_additional_info(candidate, df)
                }
                requirements.append(requirement)
            
            # Also include high-confidence edge cases (user can decide)
            edge_case_requirements = []
            for candidate in validation_results['edge_cases']:
                if candidate.confidence_score >= 0.5:  # High confidence edge cases
                    requirement = {
                        'description': candidate.content,
                        'source': f"{sheet_name}!{candidate.source_column}{candidate.row_index + 1}",
                        'sheet_name': sheet_name,
                        'row_number': candidate.metadata['original_row'] + 1,
                        'column_name': candidate.source_column,
                        'original_id': self._extract_id_from_metadata(candidate, df),
                        'confidence_score': candidate.confidence_score,
                        'category': f"edge_case_{candidate.category}",
                        'additional_info': self._extract_additional_info(candidate, df),
                        'is_edge_case': True
                    }
                    edge_case_requirements.append(requirement)
            
            # Log comprehensive results
            total_found = len(requirements) + len(edge_case_requirements)
            self.logger.info(f"✅ Enhanced extraction complete:")
            self.logger.info(f"   📋 Validated requirements: {len(requirements)}")
            self.logger.info(f"   🔍 High-confidence edge cases: {len(edge_case_requirements)}")
            self.logger.info(f"   📊 Total potential requirements: {total_found}")
            
            # Log column contributions
            column_stats = validation_results['statistics']['column_contributions']
            for col_name, count in column_stats.items():
                if count > 0:
                    self.logger.info(f"     • {col_name}: {count} requirements")
            
            # Include edge cases in final result (let downstream decide)
            final_requirements = requirements + edge_case_requirements
            
            return final_requirements
            
        except Exception as e:
            self.logger.error(f"Error extracting requirements from sheet '{sheet_name}': {str(e)}")
            return []
    
    def _extract_id_from_metadata(self, candidate, df: pd.DataFrame) -> str:
        """Extract ID from the same row if available"""
        try:
            original_row = candidate.metadata['original_row']
            # Look for ID in common ID columns
            for col in df.columns:
                col_lower = col.lower()
                if any(pattern in col_lower for pattern in ['id', 'no', 'number', 'ref']):
                    if original_row < len(df) and not pd.isna(df.iloc[original_row][col]):
                        return str(df.iloc[original_row][col]).strip()
            return ''
        except Exception:
            return ''
    
    def _extract_additional_info(self, candidate, df: pd.DataFrame) -> Dict[str, Any]:
        """Extract additional information from the same row"""
        try:
            original_row = candidate.metadata['original_row']
            additional_info = {}
            
            if original_row < len(df):
                row_data = df.iloc[original_row]
                for col_name, value in row_data.items():
                    if col_name != candidate.source_column and not pd.isna(value):
                        value_str = str(value).strip()
                        if len(value_str) > 0:
                            additional_info[col_name] = value_str
                            
            # Add validation metadata
            additional_info.update({
                'word_count': candidate.metadata['word_count'],
                'has_numbers': candidate.metadata['has_numbers'],
                'has_special_chars': candidate.metadata['has_special_chars'],
                'language_detected': candidate.metadata['language_detected']
            })
            
            return additional_info
        except Exception:
            return {}
    
    def get_lightweight_requirement_count(self, sheet_name: str, file_info: Dict) -> int:
        """Fast requirement counting without full extraction"""
        try:
            if sheet_name not in file_info['pandas_sheets']:
                return 0
            
            df = file_info['pandas_sheets'][sheet_name]
            return self.validator.get_lightweight_count(df, sheet_name)
            
        except Exception as e:
            self.logger.error(f"Error counting requirements in sheet '{sheet_name}': {str(e)}")
            return 0
            id_column = self._select_id_column(sheet_analysis['detected_columns'])
            
            # Extract requirements
            requirements = []
            
            for row_idx, row in df.iterrows():
                # Skip empty requirement rows
                if pd.isna(row[primary_req_column]) or str(row[primary_req_column]).strip() == '':
                    continue
                
                # Extract requirement data
                requirement = {
                    'description': str(row[primary_req_column]).strip(),
                    'source': f"{sheet_name}!{get_column_letter(df.columns.get_loc(primary_req_column) + 1)}{row_idx + 2}",
                    'sheet_name': sheet_name,
                    'row_number': row_idx + 2,  # Excel row number (1-indexed + header)
                    'original_id': '',
                    'additional_info': {}
                }
                
                # Add original ID if exists
                if id_column and not pd.isna(row[id_column]):
                    requirement['original_id'] = str(row[id_column]).strip()
                
                # Add other column data as additional info
                for col_name in df.columns:
                    if col_name not in [primary_req_column, id_column] and not pd.isna(row[col_name]):
                        requirement['additional_info'][col_name] = str(row[col_name]).strip()
                
                # Handle merged cells content
                merged_content = self._extract_merged_cell_content(
                    worksheet, row_idx + 2, sheet_analysis['merged_cells']
                )
                if merged_content:
                    requirement['merged_cell_content'] = merged_content
                
                requirements.append(requirement)
            
            self.logger.info(f"✅ Extracted {len(requirements)} requirements from '{sheet_name}'")
            
            return requirements
            
        except Exception as e:
            self.logger.error(f"Error extracting requirements from sheet '{sheet_name}': {str(e)}")
            return []
    
    def _select_primary_requirement_column(self, requirement_columns: List[str], 
                                         column_analysis: Dict) -> str:
        """
        Select the best requirement column based on analysis
        """
        if not requirement_columns:
            return None
        
        if len(requirement_columns) == 1:
            return requirement_columns[0]
        
        # Find column with highest confidence
        best_column = requirement_columns[0]
        best_confidence = column_analysis[best_column]['confidence']
        
        for col in requirement_columns[1:]:
            confidence = column_analysis[col]['confidence']
            if confidence > best_confidence:
                best_column = col
                best_confidence = confidence
        
        return best_column
    
    def _select_id_column(self, column_analysis: Dict) -> Optional[str]:
        """
        Select the best ID column
        """
        id_columns = [col for col, data in column_analysis.items() if data['is_id_column']]
        
        if not id_columns:
            return None
        
        if len(id_columns) == 1:
            return id_columns[0]
        
        # Prefer column with highest confidence
        best_column = id_columns[0]
        best_confidence = column_analysis[best_column]['confidence']
        
        for col in id_columns[1:]:
            confidence = column_analysis[col]['confidence']
            if confidence > best_confidence:
                best_column = col
                best_confidence = confidence
        
        return best_column
    
    def _extract_merged_cell_content(self, worksheet, row_number: int, 
                                   merged_cells: List[Dict]) -> Dict:
        """
        Extract content from merged cells for a specific row
        """
        merged_content = {}
        
        for merged_cell in merged_cells:
            if (merged_cell['start_row'] <= row_number <= merged_cell['end_row'] and 
                merged_cell['value'] is not None):
                merged_content[merged_cell['range']] = str(merged_cell['value'])
        
        return merged_content
    
    def get_sheet_suggestions_for_focus(self, file_info: Dict) -> List[Dict]:
        """
        Suggest sheets that would be good candidates for focus analysis
        """
        suggestions = []
        
        for sheet_name, analysis in file_info['sheets_analysis'].items():
            if analysis.get('has_requirements', False):
                suggestion = {
                    'sheet_name': sheet_name,
                    'confidence_score': analysis['confidence_score'],
                    'total_requirements': len(analysis.get('potential_requirement_columns', [])),
                    'total_rows': analysis.get('total_rows', 0),
                    'recommendation_reason': self._get_recommendation_reason(analysis)
                }
                suggestions.append(suggestion)
        
        # Sort by confidence score (descending)
        suggestions.sort(key=lambda x: x['confidence_score'], reverse=True)
        
        return suggestions
    
    def _get_recommendation_reason(self, analysis: Dict) -> str:
        """
        Generate human-readable reason for recommendation
        """
        reasons = []
        
        if analysis['confidence_score'] > 0.8:
            reasons.append("High confidence requirement detection")
        
        req_cols = len(analysis.get('potential_requirement_columns', []))
        if req_cols > 1:
            reasons.append(f"Multiple requirement columns ({req_cols})")
        
        if analysis.get('total_rows', 0) > 20:
            reasons.append("Substantial content volume")
        
        id_cols = sum(1 for col_data in analysis.get('detected_columns', {}).values() 
                     if col_data.get('is_id_column', False))
        if id_cols > 0:
            reasons.append("Contains ID columns")
        
        return "; ".join(reasons) if reasons else "Contains identifiable requirements"