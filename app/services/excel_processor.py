import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from typing import Dict, List, Any, Optional
from pathlib import Path
import re

from app.config import settings
from app.utils.logger import get_logger
from app.utils.exceptions import ExcelProcessingError
from app.models.requirement import Requirement, RequirementType, Priority, Status

logger = get_logger(__name__)

class ExcelProcessor:
    def __init__(self):
        self.logger = logger
    
    async def read_excel_file(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Read all sheets from Excel file
        Returns: Dictionary with sheet names as keys, DataFrames as values
        """
        try:
            self.logger.info(f"Reading Excel file: {file_path}")
            
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            sheets_data = {}
            
            for sheet_name in excel_file.sheet_names:
                self.logger.debug(f"Reading sheet: {sheet_name}")
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                sheets_data[sheet_name] = df
                
            self.logger.info(f"Successfully read {len(sheets_data)} sheets")
            return sheets_data
            
        except Exception as e:
            self.logger.error(f"Error reading Excel file: {str(e)}")
            raise ExcelProcessingError(f"Failed to read Excel file: {str(e)}")
    
    def extract_requirements_from_sheet(self, df: pd.DataFrame, sheet_name: str) -> List[Dict]:
        """
        Extract requirements from a specific sheet
        Focus on identifying text that looks like requirements
        """
        try:
            self.logger.info(f"Extracting requirements from sheet: {sheet_name}")
            requirements = []
            
            # Convert all cells to string and look for requirement-like content
            for row_idx, row in df.iterrows():
                for col_idx, cell_value in enumerate(row):
                    if pd.notna(cell_value) and isinstance(cell_value, str):
                        cell_text = str(cell_value).strip()
                        
                        # Skip empty cells or cells with only numbers/dates
                        if len(cell_text) < 10:
                            continue
                            
                        # Look for requirement-like patterns
                        if self._is_likely_requirement(cell_text):
                            requirement = {
                                'description': cell_text,
                                'source': f"{sheet_name}!{self._get_excel_column_name(col_idx)}{row_idx + 1}",
                                'sheet_name': sheet_name,
                                'row': row_idx + 1,
                                'column': col_idx + 1
                            }
                            requirements.append(requirement)
            
            self.logger.info(f"Found {len(requirements)} potential requirements in {sheet_name}")
            return requirements
            
        except Exception as e:
            self.logger.error(f"Error extracting requirements from {sheet_name}: {str(e)}")
            raise ExcelProcessingError(f"Failed to extract requirements from {sheet_name}: {str(e)}")
    
    def _is_likely_requirement(self, text: str) -> bool:
        """
        Determine if text looks like a requirement
        """
        # Skip headers, labels, and metadata
        skip_patterns = [
            r'^(requirement|description|id|type|priority|status)',
            r'^(sheet|tab|page|section)',
            r'^\d+\.\d+\.\d+',  # version numbers
            r'^(created|modified|updated|date)',
            r'^(author|owner|responsible)',
        ]
        
        text_lower = text.lower()
        for pattern in skip_patterns:
            if re.match(pattern, text_lower):
                return False
        
        # Look for requirement indicators
        requirement_indicators = [
            'shall', 'must', 'should', 'will', 'can', 'may',
            'system', 'user', 'application', 'interface',
            'function', 'feature', 'capability', 'requirement',
            'able to', 'needs to', 'required to', 'designed to'
        ]
        
        # Text should be substantial and contain requirement language
        if len(text) > 20 and any(indicator in text_lower for indicator in requirement_indicators):
            return True
            
        return False
    
    def _get_excel_column_name(self, col_index: int) -> str:
        """Convert column index to Excel column name (A, B, C, ..., AA, AB, etc.)"""
        result = ""
        while col_index >= 0:
            result = chr(col_index % 26 + 65) + result
            col_index = col_index // 26 - 1
            if col_index < 0:
                break
        return result
    
    def identify_requirement_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        """
        Auto-detect columns containing requirements
        Returns: Mapping of expected fields to actual column names
        """
        try:
            column_mapping = {}
            
            # Look for header row
            for row_idx in range(min(5, len(df))):  # Check first 5 rows
                row = df.iloc[row_idx]
                for col_idx, cell_value in enumerate(row):
                    if pd.notna(cell_value):
                        cell_text = str(cell_value).lower().strip()
                        
                        # Map common column names
                        if any(keyword in cell_text for keyword in ['requirement', 'description', 'spec']):
                            column_mapping['description'] = col_idx
                        elif any(keyword in cell_text for keyword in ['id', 'identifier', 'number']):
                            column_mapping['id'] = col_idx
                        elif any(keyword in cell_text for keyword in ['type', 'category']):
                            column_mapping['type'] = col_idx
                        elif any(keyword in cell_text for keyword in ['priority', 'importance']):
                            column_mapping['priority'] = col_idx
                        elif any(keyword in cell_text for keyword in ['status', 'state']):
                            column_mapping['status'] = col_idx
            
            return column_mapping
            
        except Exception as e:
            self.logger.error(f"Error identifying requirement columns: {str(e)}")
            return {}
    
    def generate_rtm_excel(self, requirements: List[Requirement], output_path: str) -> str:
        """
        Create formatted Excel RTM file with 3 sheets:
        1. Tool Requirements Focus
        2. Complete Requirements Matrix  
        3. Summary Statistics
        """
        try:
            self.logger.info(f"Generating RTM Excel file: {output_path}")
            
            # Separate tool requirements from all requirements
            tool_requirements = [req for req in requirements if "tool Requirements" in req.source]
            
            self.logger.info(f"Found {len(tool_requirements)} tool requirements out of {len(requirements)} total")
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Sheet 1: Tool Requirements Focus
            ws_tool = wb.active
            ws_tool.title = "Tool Requirements Focus"
            self._create_requirements_sheet(ws_tool, tool_requirements, "Tool Requirements (Priority Focus)")
            
            # Sheet 2: Complete Requirements Matrix
            ws_complete = wb.create_sheet("Complete Requirements Matrix")
            self._create_requirements_sheet(ws_complete, requirements, "All Requirements Traceability Matrix")
            
            # Sheet 3: Summary Statistics
            self._create_summary_sheet(wb, requirements, tool_requirements)
            
            # Save file
            wb.save(output_path)
            self.logger.info(f"RTM Excel file saved with 3 sheets: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error generating RTM Excel: {str(e)}")
            raise ExcelProcessingError(f"Failed to generate RTM Excel: {str(e)}")
    
    def _create_requirements_sheet(self, ws, requirements: List[Requirement], sheet_title: str):
        """Create a requirements sheet with proper formatting"""
        # Define headers
        headers = [
            "Requirement ID",
            "Requirement Description", 
            "Source",
            "Requirement Type",
            "Priority",
            "Status",
            "Related Deliverables",
            "Test Case ID",
            "Comments"
        ]
        
        # Set headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(cell)
        
        # Set column widths
        column_widths = [15, 50, 20, 15, 12, 15, 25, 15, 30]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[self._get_excel_column_name(col_idx-1)].width = width
        
        # Add data
        for row_idx, req in enumerate(requirements, 2):
            ws.cell(row=row_idx, column=1, value=req.id)
            ws.cell(row=row_idx, column=2, value=req.description)
            ws.cell(row=row_idx, column=3, value=req.source)
            ws.cell(row=row_idx, column=4, value=req.requirement_type.value if hasattr(req.requirement_type, 'value') else str(req.requirement_type))
            ws.cell(row=row_idx, column=5, value=req.priority.value if hasattr(req.priority, 'value') else str(req.priority))
            ws.cell(row=row_idx, column=6, value=req.status.value if hasattr(req.status, 'value') else str(req.status))
            ws.cell(row=row_idx, column=7, value=req.related_deliverables or "")
            ws.cell(row=row_idx, column=8, value=req.test_case_id)
            ws.cell(row=row_idx, column=9, value=req.comments or "")
            
            # Apply wrap text for description and comments
            ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row_idx, column=9).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Add data validation
        self._add_data_validation(ws, len(requirements))
    
    def _apply_header_style(self, cell):
        """Apply header styling"""
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _add_data_validation(self, ws, num_requirements: int):
        """Add dropdown validations for specific columns"""
        # Requirement Type validation
        req_type_validation = DataValidation(
            type="list",
            formula1='"Functional,Non-functional,Business,Technical,User"'
        )
        ws.add_data_validation(req_type_validation)
        req_type_validation.add(f"D2:D{num_requirements + 1}")
        
        # Priority validation
        priority_validation = DataValidation(
            type="list",
            formula1='"High,Medium,Low"'
        )
        ws.add_data_validation(priority_validation)
        priority_validation.add(f"E2:E{num_requirements + 1}")
        
        # Status validation
        status_validation = DataValidation(
            type="list", 
            formula1='"Not Tested,In Progress,Approved,Rejected"'
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f"F2:F{num_requirements + 1}")
    
    def _create_summary_sheet(self, wb, all_requirements: List[Requirement], tool_requirements: List[Requirement]):
        """Create enhanced summary statistics sheet"""
        ws_summary = wb.create_sheet("Summary Statistics")
        
        # Helper function to safely get enum values
        def safe_get_value(attr):
            return attr.value if hasattr(attr, 'value') else str(attr)
        
        # Count statistics for all requirements
        all_type_counts = {}
        all_priority_counts = {}
        all_status_counts = {}
        
        for req in all_requirements:
            req_type = safe_get_value(req.requirement_type)
            priority = safe_get_value(req.priority)
            status = safe_get_value(req.status)
            
            all_type_counts[req_type] = all_type_counts.get(req_type, 0) + 1
            all_priority_counts[priority] = all_priority_counts.get(priority, 0) + 1
            all_status_counts[status] = all_status_counts.get(status, 0) + 1
        
        # Count statistics for tool requirements
        tool_type_counts = {}
        tool_priority_counts = {}
        
        for req in tool_requirements:
            req_type = safe_get_value(req.requirement_type)
            priority = safe_get_value(req.priority)
            
            tool_type_counts[req_type] = tool_type_counts.get(req_type, 0) + 1
            tool_priority_counts[priority] = tool_priority_counts.get(priority, 0) + 1
        
        # Create summary layout
        row = 1
        
        # Overall Summary
        ws_summary.cell(row=row, column=1, value="📊 Requirements Analysis Summary")
        self._apply_header_style(ws_summary.cell(row=row, column=1))
        ws_summary.merge_cells(f'A{row}:D{row}')
        
        row += 2
        ws_summary.cell(row=row, column=1, value="Total Requirements:")
        ws_summary.cell(row=row, column=2, value=len(all_requirements))
        row += 1
        ws_summary.cell(row=row, column=1, value="Tool Requirements:")
        ws_summary.cell(row=row, column=2, value=len(tool_requirements))
        row += 1
        ws_summary.cell(row=row, column=1, value="Other Requirements:")
        ws_summary.cell(row=row, column=2, value=len(all_requirements) - len(tool_requirements))
        
        # All Requirements Breakdown
        row += 3
        ws_summary.cell(row=row, column=1, value="🔍 All Requirements Breakdown")
        self._apply_header_style(ws_summary.cell(row=row, column=1))
        
        row += 2
        ws_summary.cell(row=row, column=1, value="By Type:")
        self._apply_header_style(ws_summary.cell(row=row, column=1))
        row += 1
        for req_type, count in all_type_counts.items():
            ws_summary.cell(row=row, column=1, value=req_type)
            ws_summary.cell(row=row, column=2, value=count)
            percentage = (count / len(all_requirements)) * 100
            ws_summary.cell(row=row, column=3, value=f"{percentage:.1f}%")
            row += 1
        
        row += 1
        ws_summary.cell(row=row, column=1, value="By Priority:")
        self._apply_header_style(ws_summary.cell(row=row, column=1))
        row += 1
        for priority, count in all_priority_counts.items():
            ws_summary.cell(row=row, column=1, value=priority)
            ws_summary.cell(row=row, column=2, value=count)
            percentage = (count / len(all_requirements)) * 100
            ws_summary.cell(row=row, column=3, value=f"{percentage:.1f}%")
            row += 1
        
        # Tool Requirements Focus
        if tool_requirements:
            row += 3
            ws_summary.cell(row=row, column=1, value="⚙️ Tool Requirements Focus")
            self._apply_header_style(ws_summary.cell(row=row, column=1))
            
            row += 2
            ws_summary.cell(row=row, column=1, value="By Type:")
            self._apply_header_style(ws_summary.cell(row=row, column=1))
            row += 1
            for req_type, count in tool_type_counts.items():
                ws_summary.cell(row=row, column=1, value=req_type)
                ws_summary.cell(row=row, column=2, value=count)
                percentage = (count / len(tool_requirements)) * 100
                ws_summary.cell(row=row, column=3, value=f"{percentage:.1f}%")
                row += 1
            
            row += 1
            ws_summary.cell(row=row, column=1, value="By Priority:")
            self._apply_header_style(ws_summary.cell(row=row, column=1))
            row += 1
            for priority, count in tool_priority_counts.items():
                ws_summary.cell(row=row, column=1, value=priority)
                ws_summary.cell(row=row, column=2, value=count)
                percentage = (count / len(tool_requirements)) * 100
                ws_summary.cell(row=row, column=3, value=f"{percentage:.1f}%")
                row += 1
        
        # Set column widths for summary sheet
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15  
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 15
