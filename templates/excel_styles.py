"""
Excel styling definitions for RTM generation
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Color palette
COLORS = {
    'header_bg': '366092',      # Dark blue
    'header_text': 'FFFFFF',    # White
    'high_priority': 'FF6B6B',  # Red
    'medium_priority': 'FFE66D', # Yellow
    'low_priority': '4ECDC4',   # Green
    'border': '000000',         # Black
    'alt_row': 'F8F9FA'        # Light gray
}

# Font definitions
HEADER_FONT = Font(
    name='Calibri',
    size=12,
    bold=True,
    color=COLORS['header_text']
)

DATA_FONT = Font(
    name='Calibri',
    size=11,
    color='000000'
)

# Fill definitions
HEADER_FILL = PatternFill(
    start_color=COLORS['header_bg'],
    end_color=COLORS['header_bg'],
    fill_type='solid'
)

ALT_ROW_FILL = PatternFill(
    start_color=COLORS['alt_row'],
    end_color=COLORS['alt_row'],
    fill_type='solid'
)

# Alignment definitions
HEADER_ALIGNMENT = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True
)

DATA_ALIGNMENT = Alignment(
    horizontal='left',
    vertical='top',
    wrap_text=True
)

CENTER_ALIGNMENT = Alignment(
    horizontal='center',
    vertical='center'
)

# Border definitions
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Data validation definitions
REQUIREMENT_TYPE_VALIDATION = DataValidation(
    type="list",
    formula1='"Functional,Non-functional,Business,Technical,User"',
    allow_blank=False
)

PRIORITY_VALIDATION = DataValidation(
    type="list",
    formula1='"High,Medium,Low"',
    allow_blank=False
)

STATUS_VALIDATION = DataValidation(
    type="list",
    formula1='"Not Tested,In Progress,Approved,Rejected"',
    allow_blank=False
)

# Column configurations
RTM_COLUMNS = [
    {
        'name': 'Requirement ID',
        'width': 15,
        'alignment': CENTER_ALIGNMENT,
        'font': DATA_FONT
    },
    {
        'name': 'Requirement Description',
        'width': 50,
        'alignment': DATA_ALIGNMENT,
        'font': DATA_FONT
    },
    {
        'name': 'Source',
        'width': 20,
        'alignment': CENTER_ALIGNMENT,
        'font': DATA_FONT
    },
    {
        'name': 'Requirement Type',
        'width': 15,
        'alignment': CENTER_ALIGNMENT,
        'font': DATA_FONT,
        'validation': REQUIREMENT_TYPE_VALIDATION
    },
    {
        'name': 'Priority',
        'width': 12,
        'alignment': CENTER_ALIGNMENT,
        'font': DATA_FONT,
        'validation': PRIORITY_VALIDATION
    },
    {
        'name': 'Status',
        'width': 15,
        'alignment': CENTER_ALIGNMENT,
        'font': DATA_FONT,
        'validation': STATUS_VALIDATION
    },
    {
        'name': 'Related Deliverables',
        'width': 25,
        'alignment': DATA_ALIGNMENT,
        'font': DATA_FONT
    },
    {
        'name': 'Test Case ID',
        'width': 15,
        'alignment': CENTER_ALIGNMENT,
        'font': DATA_FONT
    },
    {
        'name': 'Comments',
        'width': 30,
        'alignment': DATA_ALIGNMENT,
        'font': DATA_FONT
    }
]

def apply_rtm_styling(worksheet, num_requirements: int):
    """Apply complete styling to RTM worksheet"""
    
    # Apply header styling
    for col_idx, col_config in enumerate(RTM_COLUMNS, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        
        # Set column width
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        worksheet.column_dimensions[column_letter].width = col_config['width']
    
    # Apply data styling
    for row_idx in range(2, num_requirements + 2):
        for col_idx, col_config in enumerate(RTM_COLUMNS, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = col_config['font']
            cell.alignment = col_config['alignment']
            cell.border = THIN_BORDER
            
            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
    
    # Apply data validations
    for col_idx, col_config in enumerate(RTM_COLUMNS, 1):
        if 'validation' in col_config:
            validation = col_config['validation']
            worksheet.add_data_validation(validation)
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            validation.add(f"{column_letter}2:{column_letter}{num_requirements + 1}")
    
    # Freeze panes (freeze header row)
    worksheet.freeze_panes = "A2"
